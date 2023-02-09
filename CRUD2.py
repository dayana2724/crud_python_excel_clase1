from datetime import datetime
from openpyxl import load_workbook

Rut=r"C:\Users\SENA\Desktop"

def leer(ruta:str, extraer:str):
    Archivo_Execcel = load_workbook(ruta)
    Hoja_Datos = Archivo_Execcel['Datos de crud']
    Hoja_Datos=Hoja_Datos['A2':'F'+str(Hoja_Datos.max_column)]
    
    info={}
    
    for i in Hoja_Datos:
        
        if isinstance(i[0].value, int):
            info.setdefailt(i[0].value, {'tarea':i[1].value, 'descripcion':i[2].value,
                                  'estado':i[3].value, 'fecha de inicio':i[4].value,
                                  'fecha de finalizacion':i[5].value })
    if not(extraer=='todo'):
        info=filtrar(info, extraer)
            
    for i in info: 
        print('*Tarea*')
        print('Id:'+ str(i)+'\n'+'Tilulo: '+(info[i]['tarea'])+ '\n'+'Descrpcion: ' 
              +str(info[i]['descripcion'])+ '\n'+ 'Estado:'+str(info[i]['Estado'])
              +'\n'+'Fecha creacion: '+ str(info[i]['Fecha de inicio'])
              +'\n'+'fecha de finalizacion:' +str(info[i]['fecha de finalizacion']))
       
        print()
        
    return 